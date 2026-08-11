#!/usr/bin/env python3
"""Full live re-scrape of tungsten inventory sources — 2026-08-11."""
from __future__ import annotations

import json
import os
import re
import time
import urllib.error
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from fractions import Fraction
from pathlib import Path
from statistics import median
from typing import Any

from datetime import datetime, timezone, timedelta

# Repo-root relative paths (GitHub Actions + local update.sh)
_REPO = Path(__file__).resolve().parent.parent
_CST = timezone(timedelta(hours=8))

def _now_cst() -> datetime:
    return datetime.now(_CST)

AS_OF = _now_cst().strftime("%Y-%m-%d")
UPDATED_AT = _now_cst().replace(microsecond=0).isoformat()
FX = 6.75
OZ_G = 28.3495
MAX_OZ = 0.5  # through 1/2 oz
MAX_G = MAX_OZ * OZ_G + 0.01  # ~14.185
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
OUT = Path(os.environ.get("SCRAPE_OUT", str(_REPO / "scrape-out")))
SHARE = Path(os.environ.get("SHARE_DIR", str(_REPO)))
EXCEL = Path(os.environ["EXCEL_PATH"]) if os.environ.get("EXCEL_PATH") else None
CANVAS = Path(os.environ["CANVAS_PATH"]) if os.environ.get("CANVAS_PATH") else None
SKIP_EXCEL = os.environ.get("SKIP_EXCEL", "1") != "0"
SKIP_CANVAS = os.environ.get("SKIP_CANVAS", "1") != "0"

SIZE_ORDER = [
    "1/64 oz",
    "1/32 oz",
    "3/64 oz",
    "1/16 oz",
    "3/32 oz",
    "1/8 oz",
    "3/16 oz",
    "1/4 oz",
    "5/16 oz",
    "3/8 oz",
    "1/2 oz",
]
SIZE_FRAC = {
    "1/64": 1 / 64,
    "3/64": 3 / 64,
    "1/32": 1 / 32,
    "1/16": 1 / 16,
    "3/32": 3 / 32,
    "1/8": 1 / 8,
    "3/16": 3 / 16,
    "1/4": 1 / 4,
    "5/16": 5 / 16,
    "3/8": 3 / 8,
    "1/2": 1 / 2,
}

STATUS: dict[str, Any] = {"ok": {}, "fail": {}, "notes": []}


@dataclass
class RawVariant:
    brand: str
    channel: str
    cat: str
    size: str
    oz: float
    qty: int
    pack_usd: float
    in_stock: bool
    color: str
    title: str
    url: str
    source: str
    sku: str = ""


@dataclass
class Row:
    brand: str
    channel: str
    cat: str
    size: str
    g: float
    qty: int
    packUsd: float
    unitCny: float
    yg: float
    url: str
    color: str = ""
    title: str = ""
    source: str = ""


def fetch(url: str, timeout: int = 35) -> tuple[int, str]:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            code = resp.getcode() or 200
            charset = resp.headers.get_content_charset() or "utf-8"
            return code, raw.decode(charset, "replace")
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", "replace")
        except Exception:
            body = ""
        return e.code, body
    except Exception as e:
        return -1, str(e)


def fetch_json(url: str) -> tuple[int, Any]:
    code, text = fetch(url)
    if code != 200:
        return code, None
    try:
        return code, json.loads(text)
    except Exception:
        return code, None


def parse_size(text: str) -> tuple[str | None, float | None]:
    """Parse oz size; reject compound sizes like 1 1/4, 1 1/2, 2 1/2."""
    # Exclude mixed numbers: "1 1/4oz", "1-1/2 oz", "2 1/2"
    # Note: fraction may be glued to "oz" so do not require \b after the fraction alone.
    cleaned = re.sub(
        r"(?<![0-9/])\d+\s*[\s-]\s*(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)(?=\s*oz\b|\b)",
        " EXCL_COMPOUND ",
        text,
        flags=re.I,
    )
    m = re.search(
        r"(?<![0-9.])(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)\s*(?:oz)?\b",
        cleaned,
        re.I,
    )
    if not m:
        return None, None
    key = m.group(1)
    oz = SIZE_FRAC[key]
    if oz > MAX_OZ + 1e-9:
        return None, None
    return f"{key} oz", oz


def parse_qty(text: str, default: int | None = None) -> int | None:
    patterns = [
        r"(\d+)\s*(?:per\s*pack|pk|pack|pcs|pc|count)",
        r"\(\s*(\d+)\s*per\s*pack\s*\)",
        r"/\s*(\d+)\s*pk",
        r"(\d+)\s*pk",
        r"-\s*(\d+)\s*pack",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return int(m.group(1))
    return default


def _size_key_from_text(text: str) -> str | None:
    """Normalize a size mention to '1/8' style key (no oz)."""
    size, _oz = parse_size(text)
    if not size:
        return None
    return size.replace(" oz", "")


def parse_tw_page_qty_map(html: str) -> dict[str, int]:
    """Build size→pack-qty from TW page (styleitem / data-gtm / desc table / Npk lines).

    Titles sometimes omit pk (e.g. 'Evolution … Black 1/8'); qty still appears as
    '1/8oz - 4pk' styleitem, 'Black/1/8oz/4pk' gtm styles, Weight|Quantity tables,
    or '4pk - 1/8oz, 3/16oz' description lines.
    """
    qty_map: dict[str, int] = {}

    def put(size_key: str | None, qty: int | None) -> None:
        if not size_key or not qty or qty < 1:
            return
        # keep first / prefer larger only if unset — don't overwrite with weaker
        if size_key not in qty_map:
            qty_map[size_key] = qty

    # styleitem: "1/8oz - 4pk" / "1/4oz - 2pk"
    for m in re.finditer(
        r"(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)\s*oz\s*[-–—]\s*(\d+)\s*pk",
        html,
        re.I,
    ):
        put(m.group(1), int(m.group(2)))

    # data-gtm_detail_styles: Black/1/4oz/2pk or Black Blue Flake/1/2oz/2pk
    gtm = re.search(r'data-gtm_detail_styles="([^"]+)"', html)
    if gtm:
        for part in gtm.group(1).split(","):
            part = part.strip()
            m = re.search(
                r"(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)\s*oz\s*/\s*(\d+)\s*pk",
                part,
                re.I,
            )
            if m:
                put(m.group(1), int(m.group(2)))
            else:
                # sometimes trailing /Npk without oz glued: already covered
                m2 = re.search(
                    r"(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)\s*/\s*(\d+)\s*pk",
                    part,
                    re.I,
                )
                if m2:
                    put(m2.group(1), int(m2.group(2)))

    # Description Weight|Quantity HTML tables (Gamakatsu etc.)
    for table in re.finditer(r"<table[\s\S]{0,4000}?</table>", html, re.I):
        t = table.group(0)
        if not re.search(r"Quantity|Qty|Pack", t, re.I):
            continue
        # strip tags to cells roughly via row pairs
        for row in re.finditer(r"<tr[^>]*>([\s\S]*?)</tr>", t, re.I):
            cells = re.findall(r"<t[dh][^>]*>([\s\S]*?)</t[dh]>", row.group(1), re.I)
            texts = [re.sub(r"<[^>]+>", " ", c) for c in cells]
            texts = [re.sub(r"\s+", " ", x).strip() for x in texts]
            if len(texts) < 2:
                continue
            sk = _size_key_from_text(texts[0])
            qm = re.search(r"^(\d+)$", texts[1])
            if sk and qm:
                put(sk, int(qm.group(1)))

    # Description lines: "4pk - 1/8oz, 3/16oz" / "3pk - 1/4oz, 5/16oz, 3/8oz"
    # Strip tags so <br /> between sizes does not break the chunk.
    desc = re.sub(r"<br\s*/?>", ",", html, flags=re.I)
    desc = re.sub(r"<[^>]+>", " ", desc)
    for m in re.finditer(
        r"(\d+)\s*pk\s*[-–—:]\s*([^;]*?)(?=(?:\d+\s*pk\s*[-–—:])|$)",
        desc,
        re.I,
    ):
        qty = int(m.group(1))
        chunk = m.group(2)
        # drop compound sizes inside the chunk (1-1/4oz etc.)
        chunk = re.sub(
            r"(?<![0-9/])\d+\s*[\s-]\s*(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)(?=\s*oz\b|\b)",
            " ",
            chunk,
            flags=re.I,
        )
        for sm in re.finditer(
            r"(?<![0-9.])(1/64|1/32|3/64|1/16|3/32|1/8|3/16|1/4|5/16|3/8|1/2)\s*(?:oz)?\b",
            chunk,
            re.I,
        ):
            put(sm.group(1), qty)

    return qty_map


def epic_qty_for_size(oz: float, cat: str) -> int | None:
    """Epic Baits pack qty by size from official product copy."""
    if cat == "Worm":
        if abs(oz - 1/16) < 1e-9:
            return 5
        if oz <= 3/16 + 1e-9:
            return 4
        if oz <= 3/8 + 1e-9:
            return 3
        if abs(oz - 1/2) < 1e-9:
            return 2
        return 1
    if cat == "Flipping":
        if oz < 1/4 - 1e-9:
            return None
        if oz <= 3/8 + 1e-9:
            return 3
        if abs(oz - 1/2) < 1e-9:
            return 2
        return 1
    return None


def strike_king_qty(oz: float) -> int | None:
    # TW Tour Grade / DT packaging (from product copy)
    # 4pk - 1/8, 3/16; 3pk - 1/4, 5/16, 3/8; 2pk - 1/2; 1pk - heavier
    if oz <= 3 / 16 + 1e-9:
        return 4
    if oz <= 3 / 8 + 1e-9:
        return 3
    if abs(oz - 1 / 2) < 1e-9:
        return 2
    return None


def gamakatsu_worm_qty(oz: float) -> int | None:
    # Official / TW Weight|Quantity table
    if oz <= 1 / 4 + 1e-9:
        return 4
    if abs(oz - 3 / 8) < 1e-9:
        return 3
    if abs(oz - 1 / 2) < 1e-9:
        return 2
    return None

def color_score(color: str, title: str = "") -> int:
    """Lower is better. Prefer Black / No-Chip Black."""
    s = f"{color} {title}".lower()
    score = 100
    if re.search(r"no[-\s]?chip\s*black|never\s*chip\s*black", s):
        score = 0
    elif re.search(r"\bblack\b", s) and "green pumpkin" not in s and "blue" not in s:
        score = 1
    elif "black" in s:
        score = 2
    elif "green pumpkin" in s:
        score = 10
    elif color.strip() in ("", "default title", "default"):
        score = 5
    else:
        score = 20
    # penalize non-black colors when black exists
    if any(x in s for x in ["red", "blue", "brown", "watermelon", "candy", "purple", "chartreuse"]):
        if "black" not in s:
            score += 30
    return score


def is_tw_in_stock(avail_text: str, has_web_stock: bool) -> bool:
    a = (avail_text or "").strip()
    if not a:
        return False
    # date like 08/19 or 8/19/26 => not currently in stock
    if re.match(r"^\d{1,2}/\d{1,2}", a):
        return False
    if re.search(r"out\s*of\s*stock|sold\s*out|unavailable|notify", a, re.I):
        return False
    # quantity: 10+, 9, 08, etc.
    if re.match(r"^\d+\+?$", a):
        return True
    if has_web_stock and re.search(r"\d", a):
        return True
    return False


def classify_cat(name: str) -> str:
    n = name.lower()
    if "punch" in n:
        return "Punch"
    if "flip" in n:
        return "Flipping"
    if "nail" in n:
        return "Nail"
    if "worm" in n or "bullet" in n:
        return "Worm"
    if "drop" in n or "barrel" in n or "tear" in n:
        return "Other"
    return "Other"


def brand_from_tw_name(name: str) -> str:
    # data-gtm brand like d/Nako or from product title prefix
    known = [
        ("WOO! Tungsten Never Chip", "WOO! Never Chip"),
        ("WOO! Tungsten", "WOO!"),
        ("1st Contact", "1st Contact"),
        ("Arsenal Fishing", "Arsenal Fishing"),
        ("Bullet Weights", "Bullet Weights"),
        ("Denali", "Denali"),
        ("Eagle Claw", "Eagle Claw"),
        ("Elite Tungsten", "Elite Tungsten"),
        ("Epic Baits", "Epic Baits"),
        ("Evolution", "Evolution"),
        ("Fitzgerald", "Fitzgerald"),
        ("Flat Out", "Flat Out"),
        ("Googan Baits", "Googan Baits"),
        ("Mustad", "Mustad"),
        ("Nako", "Nako"),
        ("Omega", "Omega"),
        ("Queen Tackle", "Queen Tackle"),
        ("Strike King", "Strike King"),
        ("TMO", "TMO"),
        ("Xzone", "Xzone"),
        ("Vike", "Vike"),
        ("Hawg Tech", "Hawg Tech"),
        ("VMC", "VMC"),
        ("Gamakatsu", "Gamakatsu"),
        ("Reaction Tackle", "Reaction Tackle"),
        ("Trokar", "Trokar"),
        ("Owner", "Owner"),
        ("Reins", "Reins"),
        ("Missile", "Missile"),
        ("Picasso", "Picasso"),
        ("Tungsten", "Tungsten"),  # fallback last
    ]
    for prefix, brand in known:
        if name.startswith(prefix) or prefix.lower() in name.lower()[: len(prefix) + 5]:
            if prefix == "Tungsten":
                continue
            return brand
    # first word(s)
    return name.split(" Tungsten")[0].split(" tungsten")[0].strip() or name


def yg_calc(pack_usd: float, qty: int, oz: float) -> tuple[float, float]:
    unit_cny = pack_usd * FX / qty
    yg = pack_usd * FX / (qty * oz * OZ_G)
    return round(unit_cny, 2), round(yg, 3)


# --------------- Shopify helpers ---------------

def shopify_product_json(handle_url: str) -> dict | None:
    """Load product; merge `available` from .js when .json omits it."""
    base = handle_url.rstrip("/")
    if base.endswith(".json"):
        base = base[: -len(".json")]
    code, data = fetch_json(base + ".json")
    if code != 200 or not data:
        return None
    product = data.get("product")
    if not product:
        return None
    variants = product.get("variants") or []
    if variants and all(v.get("available") is None for v in variants):
        code_js, js = fetch_json(base + ".js")
        if code_js == 200 and isinstance(js, dict):
            by_id = {v.get("id"): v for v in (js.get("variants") or [])}
            for v in variants:
                jv = by_id.get(v.get("id"))
                if jv is not None and jv.get("available") is not None:
                    v["available"] = jv["available"]
    return product


def shopify_avail_from_ld(product_url: str) -> dict[str, bool]:
    """sku -> in_stock from JSON-LD."""
    code, html = fetch(product_url)
    if code != 200:
        return {}
    out: dict[str, bool] = {}
    for ld in re.findall(
        r'<script[^>]*type="application/ld\+json"[^>]*>(.*?)</script>', html, re.S | re.I
    ):
        try:
            d = json.loads(ld)
        except Exception:
            continue
        if not isinstance(d, dict) or d.get("@type") != "Product":
            continue
        offers = d.get("offers")
        if isinstance(offers, dict):
            offers = [offers]
        if not isinstance(offers, list):
            continue
        for o in offers:
            sku = str(o.get("sku") or "")
            avail = str(o.get("availability") or "")
            if sku:
                out[sku] = "InStock" in avail
    # also try embedded available flags near variant ids from product JSON in page
    if not out:
        # pair title-ish via products.js style: "available":true near "sku"
        for m in re.finditer(
            r'"sku"\s*:\s*"([^"]*)"[^}]{0,400}?"available"\s*:\s*(true|false)|"available"\s*:\s*(true|false)[^}]{0,400}?"sku"\s*:\s*"([^"]*)"',
            html,
            re.S,
        ):
            if m.group(1) is not None:
                out[m.group(1)] = m.group(2) == "true"
            else:
                out[m.group(4)] = m.group(3) == "true"
    return out


def scrape_shopify_product(
    url: str,
    brand: str,
    channel: str,
    cat: str | None,
    source: str,
    qty_default: int | None = None,
) -> list[RawVariant]:
    prod = shopify_product_json(url)
    if not prod:
        STATUS["fail"][source] = STATUS["fail"].get(source, []) + [f"product.json fail: {url}"]
        return []
    avail_map = shopify_avail_from_ld(url)
    # collections/products sometimes have available; individual often not
    variants = prod.get("variants") or []
    # if no avail_map, try using variant.available if present
    rows: list[RawVariant] = []
    title = prod.get("title") or ""
    cat = cat or classify_cat(title)
    for v in variants:
        vtitle = v.get("title") or ""
        full = f"{title} | {vtitle} | {v.get('option1') or ''} | {v.get('option2') or ''} | {v.get('option3') or ''}"
        size, oz = parse_size(full)
        if not size or oz is None or oz > MAX_OZ + 1e-9:
            continue
        qty = parse_qty(full, None)
        if not qty and source == "epicbaitsfishing.com":
            qty = epic_qty_for_size(oz, cat)
        if not qty and brand == "Strike King":
            qty = strike_king_qty(oz)
            if not qty:
                qty = parse_qty(full, qty_default)
        if not qty:
            qty = qty_default
        if not qty:
            continue
        price = float(v.get("price") or 0)
        if price <= 0:
            continue
        sku = str(v.get("sku") or "")
        if v.get("available") is not None:
            in_stock = bool(v.get("available"))
        elif sku and sku in avail_map:
            in_stock = avail_map[sku]
        elif avail_map and not sku:
            # fallback: if only one offer matches size somehow — skip ambiguous
            in_stock = any(avail_map.values())  # weak; prefer False if mixed
            # better: don't assume
            in_stock = False
            STATUS["notes"].append(f"no sku avail mapping: {url} {vtitle}")
        else:
            # last resort: schema had no data — mark unknown as OOS for priced rows safety
            in_stock = False
            STATUS["notes"].append(f"avail unknown -> OOS: {url} {vtitle}")
        color = v.get("option1") or ""
        rows.append(
            RawVariant(
                brand=brand,
                channel=channel,
                cat=cat,
                size=size,
                oz=oz,
                qty=qty,
                pack_usd=price,
                in_stock=in_stock,
                color=color,
                title=f"{title} | {vtitle}",
                url=url.split("?")[0],
                source=source,
                sku=sku,
            )
        )
    return rows


def scrape_woo_neverchip() -> list[RawVariant]:
    url = "https://wootungsten.com/collections/neverchip/products.json?limit=250"
    code, data = fetch_json(url)
    if code != 200 or not data:
        STATUS["fail"]["wootungsten.com"] = [f"collection products.json failed code={code}"]
        return []
    out: list[RawVariant] = []
    for p in data.get("products") or []:
        title = p.get("title") or ""
        up = title.upper()
        if "NEVER CHIP" not in up:
            continue
        if "BEAD" in up or "STOP" in up or "MIXER" in up:
            continue
        if "FLIP" not in up and "WEIGHT" not in up:
            continue
        size, oz = parse_size(title)
        if not size or oz is None or oz > MAX_OZ + 1e-9:
            continue
        qty = parse_qty(title, None)
        if not qty:
            continue
        handle = p.get("handle") or ""
        product_url = (
            f"https://wootungsten.com/products/{handle}"
            if handle
            else "https://wootungsten.com/collections/neverchip"
        )
        for v in p.get("variants") or []:
            price = float(v.get("price") or 0)
            if price <= 0:
                continue
            # collection JSON may omit available — treat missing as in-stock only if price present? safer: require True
            if v.get("available") is False:
                in_stock = False
            elif v.get("available") is True:
                in_stock = True
            else:
                in_stock = True  # collection endpoint often omits flag; page listed as available earlier
            out.append(
                RawVariant(
                    brand="WOO! Never Chip",
                    channel="官网",
                    cat="Flipping",
                    size=size,
                    oz=oz,
                    qty=qty,
                    pack_usd=price,
                    in_stock=in_stock,
                    color=v.get("option1") or "Black",
                    title=title,
                    url=product_url,
                    source="wootungsten.com",
                    sku=str(v.get("sku") or ""),
                )
            )
    STATUS["ok"]["wootungsten.com"] = len(out)
    return out


# --------------- TW ---------------

def tw_catalog_urls() -> list[tuple[str, str, str]]:
    """Return list of (url, brand_hint, name)."""
    code, html = fetch("https://www.tacklewarehouse.com/catpage-TUNGSTNWTS.html", timeout=45)
    if code != 200:
        STATUS["fail"]["tacklewarehouse.com"] = [f"category page HTTP {code}"]
        return []
    items = []
    seen = set()
    for m in re.finditer(
        r'data-gtm_impression_name="([^"]+)"[^>]*data-gtm_impression_brand="([^"]+)"[^>]*',
        html,
    ):
        pass  # order varies
    # Parse cattable cells
    for m in re.finditer(
        r'data-code="([A-Z0-9]+)"\s*data-gtm_impression_code="[^"]*"\s*data-gtm_impression_name="([^"]*)"[^>]*data-gtm_impression_brand="([^"]*)"',
        html,
    ):
        code_, name, brand_raw = m.group(1), m.group(2), m.group(3)
        # find href near this code
        # search in a window
    # Simpler: all unique descpage URLs + name from nearby
    for m in re.finditer(
        r'href="\s*(https://www\.tacklewarehouse\.com/[^"]+/descpage-([A-Z0-9]+)\.html)\s*"',
        html,
    ):
        url = m.group(1).strip()
        code_ = m.group(2)
        if url in seen or "#" in url:
            continue
        seen.add(url)
        # find name
        name_m = re.search(
            rf'data-gtm_impression_code="{code_}"[^>]*data-gtm_impression_name="([^"]+)"',
            html,
        )
        if not name_m:
            name_m = re.search(
                rf'data-gtm_impression_name="([^"]+)"[^>]*data-gtm_impression_code="{code_}"',
                html,
            )
        name = name_m.group(1) if name_m else code_
        brand = brand_from_tw_name(name)
        items.append((url, brand, name))
    # Also merge known URLs from prior dataset to not miss any
    known = [
        "https://www.tacklewarehouse.com/1st_Contact_Tungsten_Flipping_Weights/descpage-FCTFW.html",
        "https://www.tacklewarehouse.com/1st_Contact_Tungsten_Worm_Weights/descpage-FCTW.html",
        "https://www.tacklewarehouse.com/Arsenal_Fishing_Flipping_Weight_Black_Out_Tungsten/descpage-AFBOF.html",
        "https://www.tacklewarehouse.com/Bullet_Weights_Tactical_Tungsten_Flipping_Weights/descpage-BWCT.html",
        "https://www.tacklewarehouse.com/Bullet_Weights_Tungsten_Bullet_Weights/descpage-BWTBW.html",
        "https://www.tacklewarehouse.com/Denali_Kovert_Tungsten_Flipping_Weights/descpage-DENTFW.html",
        "https://www.tacklewarehouse.com/Eagle_Claw_Lazer_Sharp_Tungsten_Worm_Weight/descpage-ECLTGW.html",
        "https://www.tacklewarehouse.com/Elite_Tungsten_Flippin_Weights/descpage-ELTNGFL.html",
        "https://www.tacklewarehouse.com/Elite_Tungsten_Worm_Weights/descpage-ELTUNWM.html",
        "https://www.tacklewarehouse.com/Epic_Baits_Tungsten_Flipping_Weights/descpage-EBFW.html",
        "https://www.tacklewarehouse.com/Epic_Baits_Tungsten_Worm_Weights/descpage-EBTW.html",
        "https://www.tacklewarehouse.com/Fitzgerald_Fishing_Tungsten_Worm_Weights/descpage-FFTWW.html",
        "https://www.tacklewarehouse.com/Flat_Out_Tungsten_Skirted_Punch_Weights/descpage-FOP.html",
        "https://www.tacklewarehouse.com/Googan_Baits_Green_Series_Tungsten_Flipping_Weights/descpage-GBTT.html",
        "https://www.tacklewarehouse.com/Googan_Baits_Green_Series_Tungsten_Worm_Weights/descpage-GBTW.html",
        "https://www.tacklewarehouse.com/Mustad_Tungsten_TitanX_Worm_Weights/descpage-MTXT.html",
        "https://www.tacklewarehouse.com/Nako_Tungsten_Flipping_Weights/descpage-NTFW.html",
        "https://www.tacklewarehouse.com/Nako_Tungsten_Worm_Weights_5pk/descpage-NWW.html",
        "https://www.tacklewarehouse.com/Omega_Tungsten_Flip_Weights/descpage-OTF.html",
        "https://www.tacklewarehouse.com/Queen_Tackle_Tungsten_Flipping_Weights/descpage-JHSDFHD.html",
        "https://www.tacklewarehouse.com/Queen_Tackle_Tungsten_Worm_Weights/descpage-QTWW.html",
        "https://www.tacklewarehouse.com/Strike_King_Tour_Grade_Tungsten_Weights/descpage-SKTGT.html",
        "https://www.tacklewarehouse.com/TMO_Tungsten_Flipping_Weights/descpage-TMOTF.html",
        "https://www.tacklewarehouse.com/TMO_Tungsten_Punching_Weights/descpage-TMOTP.html",
        "https://www.tacklewarehouse.com/TMO_Tungsten_Worm_Weights/descpage-TMOTW.html",
        "https://www.tacklewarehouse.com/WOO_Tungsten_Flipping_Weights/descpage-WTFW.html",
        "https://www.tacklewarehouse.com/WOO_Tungsten_Never_Chip_Flipping_Weights/descpage-WTNC.html",
        "https://www.tacklewarehouse.com/Xzone_Tungsten_Flipping_Weights/descpage-XRTF.html",
        "https://www.tacklewarehouse.com/Xzone_Tungsten_Worm_Weights/descpage-XRTW.html",
    ]
    have = {u for u, _, _ in items}
    for u in known:
        if u not in have:
            slug = u.split("/")[-2].replace("_", " ")
            items.append((u, brand_from_tw_name(slug), slug))
    return items


def scrape_tw_product(url: str, brand_hint: str, name_hint: str) -> list[RawVariant]:
    html = ""
    for attempt in range(4):
        code, html = fetch(url, timeout=45)
        if code == 200 and "js-ordering-subproduct" in html:
            break
        time.sleep(0.8 * (attempt + 1))
    else:
        if code != 200 or "js-ordering-subproduct" not in html:
            return []
    owner = re.search(r'data-owner-name="([^"]+)"', html)
    product_name = owner.group(1) if owner else name_hint
    brand = brand_hint or brand_from_tw_name(product_name)
    # special brand overrides
    if "Never Chip" in product_name or "NEVER CHIP" in product_name:
        brand = "WOO! Never Chip"
    elif product_name.startswith("WOO!"):
        brand = "WOO!"
    cat = classify_cat(product_name)
    if "Tour Grade" in product_name and "Worm" not in product_name and "Flip" not in product_name:
        # Tour Grade tungsten weights are worm-style bullets
        cat = "Worm"
    page_qty = parse_tw_page_qty_map(html)
    out: list[RawVariant] = []
    parts = re.split(r"(?=<tr[^>]*js-ordering-subproduct)", html)
    for p in parts[1:]:
        name_m = re.search(r'itemOffered">([^<]+)', p)
        avail_m = re.search(r'js-ordering-available[^>]*>([^<]+)', p)
        price_m = re.search(r'js-ordering-price[^>]*>([^<]+)', p)
        if not name_m or not price_m:
            continue
        name = name_m.group(1).strip()
        avail = avail_m.group(1).strip() if avail_m else ""
        try:
            price = float(price_m.group(1).strip())
        except ValueError:
            continue
        has_web = bool(re.search(r"<tr[^>]{0,400}data-web-stock", p[:800], re.S))
        # notify-me rows are OOS even if qty text looks numeric
        if re.search(r"data-stock-notify", p[:900], re.S):
            in_stock = False
        else:
            in_stock = is_tw_in_stock(avail, has_web)
        # Prefer size from styleitem when title omits oz / uses ambiguous fractions
        style_sizes = re.findall(
            r'js-ordering-style-item-name styleitem">([^<]+)', p
        )
        size_probe = " ".join(style_sizes) + " " + name
        size, oz = parse_size(size_probe)
        if not size or oz is None or oz > MAX_OZ + 1e-9:
            continue
        size_key = size.replace(" oz", "")
        # qty: title/row → styleitem/gtm/desc map → brand fallbacks
        qty = parse_qty(name, None) or parse_qty(p, None) or parse_qty(product_name, None)
        if not qty:
            for st in style_sizes:
                qty = parse_qty(st, None)
                if qty:
                    break
        if not qty:
            qty = page_qty.get(size_key)
        if not qty:
            m_pk = re.search(
                rf"{re.escape(size_key)}\s*oz\s*/\s*(\d+)\s*pk|"
                rf"{re.escape(size_key)}\s*oz[^\d]{{0,12}}(\d+)\s*pk|"
                rf"{re.escape(size)}[^\d]{{0,12}}(\d+)\s*pk",
                html,
                re.I,
            )
            if m_pk:
                qty = int(next(g for g in m_pk.groups() if g))
        if not qty and brand == "Strike King":
            qty = strike_king_qty(oz)
        if not qty and brand == "Gamakatsu" and cat == "Worm":
            qty = gamakatsu_worm_qty(oz)
        if not qty:
            continue
        # color from name / styleitem
        color = ""
        cm = re.search(
            r"(No-?Chip Black|Never Chip Black|Black Out|Black Blue Flake|Blk Blue Flake|Matte Black|Black|Green Pumpkin|GP|Brown|Red|Blue|Watermelon|Candy|Junebug)",
            name,
            re.I,
        )
        if not cm:
            for st in style_sizes:
                cm = re.search(
                    r"(No-?Chip Black|Never Chip Black|Black Out|Black Blue Flake|Matte Black|Black|Green Pumpkin|Junebug|Watermelon)",
                    st,
                    re.I,
                )
                if cm:
                    break
        if cm:
            color = cm.group(1)
        out.append(
            RawVariant(
                brand=brand,
                channel="TW",
                cat=cat,
                size=size,
                oz=oz,
                qty=qty,
                pack_usd=price,
                in_stock=in_stock,
                color=color,
                title=name,
                url=url,
                source="tacklewarehouse.com",
                sku=re.search(r'data-code="([^"]+)"', p).group(1)
                if re.search(r'data-code="([^"]+)"', p)
                else "",
            )
        )
    return out


def prefer_color_rows(raw: list[RawVariant]) -> tuple[list[Row], list[RawVariant]]:
    """Select best in-stock color per brand/channel/cat/size; keep OOS for oos list."""
    groups: dict[tuple, list[RawVariant]] = defaultdict(list)
    for r in raw:
        key = (r.brand, r.channel, r.cat, r.size)
        groups[key].append(r)

    instock_rows: list[Row] = []
    oos_kept: list[RawVariant] = []

    for key, variants in groups.items():
        stocked = [v for v in variants if v.in_stock]
        if stocked:
            stocked.sort(key=lambda v: (color_score(v.color, v.title), v.pack_usd))
            best = stocked[0]
            unit, yg = yg_calc(best.pack_usd, best.qty, best.oz)
            instock_rows.append(
                Row(
                    brand=best.brand,
                    channel=best.channel,
                    cat=best.cat,
                    size=best.size,
                    g=round(best.oz * OZ_G, 3),
                    qty=best.qty,
                    packUsd=best.pack_usd,
                    unitCny=unit,
                    yg=yg,
                    url=best.url,
                    color=best.color,
                    title=best.title,
                    source=best.source,
                )
            )
            # other colors oos not needed; same size other colors that are oos irrelevant
        else:
            # all OOS — keep one representative for 缺货清单 (prefer black) without inventing as priced
            variants.sort(key=lambda v: (color_score(v.color, v.title), v.pack_usd))
            oos_kept.append(variants[0])

    return instock_rows, oos_kept


def scrape_all_sources() -> tuple[list[Row], list[RawVariant], dict]:
    all_raw: list[RawVariant] = []

    # --- Shopify brand sites ---
    shopify_jobs = [
        (
            "https://nakoshop.com/products/tungsten-worm-weights",
            "Nako",
            "官网",
            "Worm",
            "nakoshop.com",
            5,
        ),
        (
            "https://nakoshop.com/products/nako-tungsten-flipping-weights",
            "Nako",
            "官网",
            "Flipping",
            "nakoshop.com",
            5,
        ),
        (
            "https://nakoshop.com/products/tungsten-nail-weights",
            "Nako",
            "官网",
            "Nail",
            "nakoshop.com",
            10,
        ),
        (
            "https://nakoshop.com/products/tungsten-skirted-nail-weights",
            "Nako",
            "官网",
            "Nail",
            "nakoshop.com",
            5,
        ),
        (
            "https://epicbaitsfishing.com/products/tungsten-worm-weight",
            "Epic Baits",
            "官网",
            "Worm",
            "epicbaitsfishing.com",
            None,
        ),
        (
            "https://epicbaitsfishing.com/products/tungsten-flipping-weight",
            "Epic Baits",
            "官网",
            "Flipping",
            "epicbaitsfishing.com",
            None,
        ),
        (
            "https://discounttackle.com/products/departure-outdoors-tungsten-worm-weights",
            "Departure",
            "DiscountTackle",
            "Worm",
            "discounttackle.com",
            5,
        ),
        (
            "https://discounttackle.com/products/strike-king-tungsten-weights",
            "Strike King",
            "DiscountTackle",
            "Worm",
            "discounttackle.com",
            None,
        ),
        (
            "https://reactiontackle.com/products/reaction-tackle-tungsten-worm-weights-bullet-shaped-sinkers",
            "Reaction Tackle",
            "官网",
            "Worm",
            "reactiontackle.com",
            None,
        ),
    ]

    print("Scraping Shopify sites...")
    for url, brand, channel, cat, source, qd in shopify_jobs:
        try:
            rows = scrape_shopify_product(url, brand, channel, cat, source, qd)
            all_raw.extend(rows)
            STATUS["ok"][source] = STATUS["ok"].get(source, 0) + len(rows)
            print(f"  {source} {brand} {cat}: {len(rows)} variants")
        except Exception as e:
            STATUS["fail"].setdefault(source, []).append(f"{url}: {e}")
            print(f"  FAIL {url}: {e}")
        time.sleep(0.3)

    print("Scraping WOO Never Chip collection...")
    try:
        woo = scrape_woo_neverchip()
        all_raw.extend(woo)
        print(f"  woo: {len(woo)}")
    except Exception as e:
        STATUS["fail"]["wootungsten.com"] = [str(e)]
        print("  FAIL woo", e)

    print("Fetching TW catalog...")
    catalog = tw_catalog_urls()
    print(f"  TW product pages: {len(catalog)}")
    (OUT / "tw_catalog.json").write_text(json.dumps(catalog, indent=2), encoding="utf-8")

    print("Scraping TW product pages...")
    tw_raw: list[RawVariant] = []
    fails = []

    def job(item):
        url, brand, name = item
        try:
            rows = scrape_tw_product(url, brand, name)
            return url, rows, None
        except Exception as e:
            return url, [], str(e)

    with ThreadPoolExecutor(max_workers=3) as ex:
        futs = [ex.submit(job, it) for it in catalog]
        for fut in as_completed(futs):
            url, rows, err = fut.result()
            if err or not rows:
                fails.append(url)
                print(f"  TW miss/fail: {url} ({err})")
            else:
                tw_raw.extend(rows)
                print(f"  TW ok {len(rows):2d}  {url.split('/')[-1]}")

    if fails:
        print(f"Retrying {len(fails)} TW pages serially...")
        still: list[str] = []
        for url in fails:
            time.sleep(0.6)
            item = next((it for it in catalog if it[0] == url), None)
            if not item:
                still.append(url)
                continue
            _, rows, err = job(item)
            if err or not rows:
                still.append(url)
                print(f"  TW retry fail: {url}")
            else:
                tw_raw.extend(rows)
                print(f"  TW retry ok {len(rows):2d}  {url.split('/')[-1]}")
        fails = still

    all_raw.extend(tw_raw)
    if tw_raw:
        STATUS["ok"]["tacklewarehouse.com"] = len(tw_raw)
    if fails:
        STATUS["fail"]["tacklewarehouse.com_pages"] = fails

    # Epic qty defaults from pack patterns in titles if missing — already handled

    instock, oos = prefer_color_rows(all_raw)
    # sort
    instock.sort(key=lambda r: (SIZE_ORDER.index(r.size) if r.size in SIZE_ORDER else 99, r.yg, r.brand))
    return instock, oos, {"raw_count": len(all_raw), "instock": len(instock), "oos_groups": len(oos)}


def brand_stats(rows: list[Row]) -> list[dict]:
    by: dict[str, list[float]] = defaultdict(list)
    sizes: dict[str, set[str]] = defaultdict(set)
    for r in rows:
        by[r.brand].append(r.yg)
        sizes[r.brand].add(r.size)
    nako_ys = by.get("Nako") or []
    nako_med = median(nako_ys) if nako_ys else 3.141

    def tier_for(med_v: float) -> str:
        if not nako_med:
            return "差不多"
        pct = (med_v - nako_med) / nako_med * 100
        if pct < -3:
            return "更便宜"
        if pct > 3:
            return "更贵"
        return "差不多"

    stats = []
    for brand, ys in by.items():
        sz = sorted(sizes[brand], key=lambda s: SIZE_ORDER.index(s) if s in SIZE_ORDER else 99)
        short = " · ".join(s.replace(" oz", "") for s in sz)
        stats.append(
            {
                "brand": brand,
                "n": len(ys),
                "sizes": short + (f" ({len(sz)})" if sz else ""),
                "sizeList": sz,
                "min": round(min(ys), 3),
                "med": round(median(ys), 3),
                "max": round(max(ys), 3),
                "tier": tier_for(median(ys)),
            }
        )
    stats.sort(key=lambda x: x["med"])
    return stats


def write_outputs(instock: list[Row], oos: list[RawVariant]):
    global AS_OF, UPDATED_AT
    AS_OF = _now_cst().strftime("%Y-%m-%d")
    UPDATED_AT = _now_cst().replace(microsecond=0).isoformat()

    stats = brand_stats(instock)
    payload = {
        "meta": {
            "title": "Nako vs 竞品 · 钨钢 ¥/克对比",
            "asOf": AS_OF,
            "updatedAt": UPDATED_AT,
            "fx": FX,
            "ozToG": OZ_G,
            "maxPieceG": 14.2,
            "formula": "¥/g = pack USD × 6.75 ÷ (qty × oz × 28.3495)",
            "rules": [
                "仅有货（in-stock / available）",
                "单粒克重不超过约 14g（含 1/2 oz ≈14.175g；更重规格已排除）",
                "同规格多色优先 Black / No-Chip Black",
                "主排名：同品牌取更便宜渠道一条",
                "有货/没货克重 = 相对本数据集中该品牌+渠道+品类清单与 Nako 同品类对照克重（芯片：oz + 克）",
            ],
            "scrapeStatus": STATUS,
        },
        "rows": [
            {
                "brand": r.brand,
                "channel": r.channel,
                "cat": r.cat,
                "size": r.size,
                "g": r.g,
                "qty": r.qty,
                "packUsd": r.packUsd,
                "unitCny": r.unitCny,
                "yg": r.yg,
                "url": r.url,
            }
            for r in instock
        ],
        "oos": [
            {
                "brand": r.brand,
                "channel": r.channel,
                "cat": r.cat,
                "size": r.size,
                "qty": r.qty,
                "packUsd": r.pack_usd,  # listed for reference but NOT used as priced row
                "color": r.color,
                "title": r.title,
                "url": r.url,
                "note": "缺货（未计入有货比价）",
            }
            for r in oos
        ],
        "brandStats": stats,
    }
    (OUT / "scrape_result.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (SHARE / "data.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("Wrote", SHARE / "data.json", "updatedAt", UPDATED_AT)

    if SKIP_EXCEL or not EXCEL:
        print("SKIP_EXCEL: skipping workbook")
        return payload

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # ---- Excel ----
    wb = Workbook()
    # 说明
    ws = wb.active
    ws.title = "说明"
    lines = [
        f"钨钢比价 · 全渠道 live 更新 {AS_OF} · 仅有货计入 ¥/克",
        "",
        "口径：¥/g=(包装$×6.75)/(粒数×oz×28.3495)；只含 available/in-stock",
        "范围：至 1/2 oz（约 14.175g）；同规格多色优先 Black / No-Chip Black",
        "来源：nakoshop / tacklewarehouse / discounttackle / epicbaits / wootungsten / reactiontackle",
        f"抓取状态：成功源={list(STATUS.get('ok',{}).keys())}",
        f"失败/告警：{STATUS.get('fail')}",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(i, 1, line)

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    # 有货明细
    ws = wb.create_sheet("有货明细")
    headers = ["品牌", "渠道", "品类", "克重", "单粒g", "粒数", "包装$", "单粒¥", "¥/g", "颜色偏好", "来源", "产品/变体", "URL"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, r in enumerate(instock, 2):
        vals = [
            r.brand,
            r.channel,
            r.cat,
            r.size,
            r.g,
            r.qty,
            r.packUsd,
            r.unitCny,
            r.yg,
            r.color,
            r.source,
            r.title,
            r.url,
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v).border = thin

    # 缺货清单
    ws = wb.create_sheet("缺货清单")
    headers = ["品牌", "渠道", "品类", "克重", "粒数", "官网标价$(仅参考未计入比价)", "颜色", "产品/变体", "URL", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="833C0C")
        cell.font = header_font
    for i, r in enumerate(oos, 2):
        vals = [
            r.brand,
            r.channel,
            r.cat,
            r.size,
            r.qty,
            r.pack_usd,
            r.color,
            r.title,
            r.url,
            "缺货 — 价格不计入有货比价",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v).border = thin

    # 明细全量_仅有货 (compat)
    ws = wb.create_sheet("明细全量_仅有货")
    headers = ["品牌", "渠道", "品类", "克重", "单粒g", "粒数", "包装$", "单粒¥", "¥/g", "来源", "产品/变体"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, r in enumerate(instock, 2):
        vals = [r.brand, r.channel, r.cat, r.size, r.g, r.qty, r.packUsd, r.unitCny, r.yg, r.source, f"{r.title} | {r.color}"]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)

    # TW全品牌_仅有货
    ws = wb.create_sheet("TW全品牌_仅有货")
    headers = ["品牌", "品类", "克重", "粒数", "包装$", "¥/g", "变体", "产品页"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    tw_rows = [r for r in instock if r.channel == "TW"]
    for i, r in enumerate(tw_rows, 2):
        vals = [r.brand, r.cat, r.size, r.qty, r.packUsd, r.yg, r.title, "tacklewarehouse.com"]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)

    # 品牌档位结论 — 有货克重 not bare n
    ws = wb.create_sheet("品牌档位结论")
    ws.cell(1, 1, "品牌价格档位（基于有货 SKU 的 ¥/g 中位数）")
    ws.cell(2, 1, f"范围：≤1/2 oz · 仅有货 · 抓取日 {AS_OF}")
    headers = ["档位", "品牌", "有货克重", "样本数", "最低¥/g", "中位¥/g", "最高¥/g"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
    tier_map = {"更便宜": "便宜", "差不多": "差不多", "更贵": "更贵"}
    for i, s in enumerate(stats, 5):
        vals = [tier_map.get(s["tier"], s["tier"]), s["brand"], s["sizes"], s["n"], s["min"], s["med"], s["max"]]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)

    # 一眼看板 1/4 oz
    ws = wb.create_sheet("一眼看板_1_4oz")
    ws.cell(1, 1, "1/4 oz 有货 · 全渠道全品牌 ¥/克排名（越低越便宜）")
    ws.cell(2, 1, f"汇率{FX}｜抓取 {AS_OF}")
    row_i = 4
    for cat in ["Worm", "Flipping", "Punch", "Other"]:
        subset = [r for r in instock if r.size == "1/4 oz" and r.cat == cat]
        if not subset:
            continue
        subset.sort(key=lambda r: r.yg)
        ws.cell(row_i, 1, f"{cat} · 1/4 oz")
        row_i += 1
        headers = ["排名", "品牌", "渠道", "包装$", "粒数", "单粒¥", "¥/g", "变体"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row_i, c, h)
            cell.fill = header_fill
            cell.font = header_font
        row_i += 1
        for rank, r in enumerate(subset, 1):
            for c, v in enumerate([rank, r.brand, r.channel, r.packUsd, r.qty, r.unitCny, r.yg, r.title], 1):
                ws.cell(row_i, c, v)
            row_i += 1
        row_i += 1

    # matrices
    def write_matrix(sheet_name: str, cat_filter):
        ws = wb.create_sheet(sheet_name)
        tw = [r for r in instock if r.channel == "TW" and cat_filter(r.cat)]
        brands = []
        brand_med = {}
        by_b: dict[str, list] = defaultdict(list)
        for r in tw:
            by_b[r.brand].append(r)
        for b, rs in by_b.items():
            brand_med[b] = median([x.yg for x in rs])
        brands = sorted(by_b.keys(), key=lambda b: brand_med[b])
        sizes = [s for s in SIZE_ORDER if any(r.size == s for r in tw)]
        ws.cell(1, 1, sheet_name.replace("_", " ") + f" · {AS_OF}")
        ws.cell(3, 1, "克重")
        for j, b in enumerate(brands, 2):
            ws.cell(3, j, b)
        for i, size in enumerate(sizes, 4):
            ws.cell(i, 1, size)
            for j, b in enumerate(brands, 2):
                cell = next((r for r in by_b[b] if r.size == size), None)
                ws.cell(i, j, cell.yg if cell else "—")

    write_matrix("TW_Worm矩阵_¥每克", lambda c: c == "Worm")
    write_matrix("TW_Flipping矩阵_¥每克", lambda c: c in ("Flipping", "Punch"))

    wb.save(EXCEL)
    print("Excel saved", EXCEL)
    return payload


def patch_static_and_canvas(payload: dict):
    """Update embedded DATA in index.html; optionally canvas."""
    rows = payload["rows"]
    stats = payload["brandStats"]

    if not SKIP_CANVAS and CANVAS and Path(CANVAS).exists():
        text = CANVAS.read_text(encoding="utf-8")
        all_compact = [
            {
                "brand": r["brand"],
                "channel": r["channel"],
                "cat": r["cat"],
                "size": r["size"],
                "g": r["g"],
                "qty": r["qty"],
                "packUsd": r["packUsd"],
                "unitCny": r["unitCny"],
                "yg": r["yg"],
                "url": r["url"],
            }
            for r in rows
        ]
        brand_compact = [
            {
                "brand": s["brand"],
                "n": s["n"],
                "min": s["min"],
                "med": s["med"],
                "max": s["max"],
                "tier": s["tier"],
            }
            for s in stats
        ]
        text2 = re.sub(
            r"const ALL: SkuRow\[\] = \[.*?\];\n",
            "const ALL: SkuRow[] = " + json.dumps(all_compact, ensure_ascii=False, separators=(",", ":")) + ";\n",
            text,
            count=1,
            flags=re.S,
        )
        text2 = re.sub(
            r"const BRAND_STATS: BrandStat\[\] = \[.*?\];\n",
            "const BRAND_STATS: BrandStat[] = " + json.dumps(brand_compact, ensure_ascii=False, separators=(",", ":")) + ";\n",
            text2,
            count=1,
            flags=re.S,
        )
        text2 = re.sub(r'const AS_OF = "[^"]+";', f'const AS_OF = "{AS_OF}";', text2)
        if "有货SKU数" in text2:
            text2 = text2.replace("有货SKU数(n)", "有货克重").replace("有货SKU数", "有货克重")
        CANVAS.write_text(text2, encoding="utf-8")
        print("Canvas updated")
    else:
        print("SKIP_CANVAS: skipping canvas update")

    html_path = SHARE / "index.html"
    html = html_path.read_text(encoding="utf-8")
    data_obj = {
        "meta": payload["meta"],
        "rows": rows,
        "brandStats": stats,
        "oos": payload.get("oos", []),
    }
    m = re.search(r"const DATA = (\{.*?\});\n", html, re.S)
    if not m:
        raise SystemExit("DATA block not found in index.html")
    new_html = html[: m.start()] + "const DATA = " + json.dumps(data_obj, ensure_ascii=False) + ";\n" + html[m.end() :]
    new_html = new_html.replace("有货SKU数(n)", "有货克重").replace(">样本数<", ">有货克重<")
    html_path.write_text(new_html, encoding="utf-8")
    print("Static index.html updated")



def main():
    import cross_validate as cv

    OUT.mkdir(parents=True, exist_ok=True)
    STATUS["ok"].clear()
    STATUS["fail"].clear()
    STATUS["notes"].clear()
    instock, oos, summary = scrape_all_sources()
    print("SUMMARY", summary)
    print("STATUS ok", STATUS.get("ok"))
    print("STATUS fail keys", list(STATUS.get("fail", {}).keys()))

    print("Cross-validating with distinct methods (primary + alternate + formula + peer + availability)...")
    do_alt = os.environ.get("SKIP_ALTERNATE", "0") != "1"
    vr = cv.validate_rows(
        instock,
        oos,
        scrape_mod=__import__(__name__),
        do_alternate_fetch=do_alt,
        delay_s=float(os.environ.get("ALT_DELAY", "0.35")),
    )
    print(
        f"Validation: published={len(vr.published)} excluded={vr.meta.get('excluded')} "
        f"disagreements={vr.meta.get('disagreements')} methods={vr.meta.get('methodCount')}"
    )
    (OUT / "validation.json").write_text(
        json.dumps(
            {
                "meta": vr.meta,
                "needs_review": vr.needs_review[:200],
                "disagreements": vr.disagreements[:200],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    summary["validation"] = vr.meta
    payload = write_outputs(vr.published, vr.oos)
    # attach validation metadata
    payload["meta"]["validatedAt"] = UPDATED_AT
    payload["meta"]["validation"] = vr.meta
    (SHARE / "data.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    patch_static_and_canvas(payload)
    (OUT / "status.json").write_text(
        json.dumps({"summary": summary, "status": STATUS, "validation": vr.meta}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print("DONE instock", len(vr.published), "oos", len(vr.oos), "needs_review", len(vr.needs_review))


if __name__ == "__main__":
    main()
